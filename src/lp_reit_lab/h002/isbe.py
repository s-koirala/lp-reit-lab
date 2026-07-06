"""ISBE Report Card parser + within-vintage school-quality rank builder (H002).

Implements the PRE-SPECIFIED recipe in
[docs/methodology/methodology_isbe-quality-recipe_2026-07-06.md](../../../docs/methodology/methodology_isbe-quality-recipe_2026-07-06.md)
(the recipe doc is normative; section references below are to it). The public
entry point is :func:`load_school_quality`, which returns one tidy long frame
of CPS school x report-card-vintage rows with the recipe's composite
proficiency metric and its within-(vintage x level) Hazen percentile rank.

Scope notes (all per the recipe):

- **Blindness** — this module reads school-side files only
  (``data/raw/isbe_report_card/``); no Cook County sale data is ever touched
  (recipe blindness statement; Rubin 2008, doi:10.1214/08-AOAS187).
- **Field resolution** — zip-era (rc2006-rc2017) field positions are resolved
  from the official ISBE record-layout workbooks by DESCRIPTION-STRING match,
  never by hard-coded position: field numbers drift year to year (the ALL
  TESTS composite is field #588 in rc2006 but #895 in rc2014 — recipe §1.3).
  Delimited-file index = layout field number - 1 (recipe §6, verified).
- **Rank units** — the recipe §1.2 pins the Hazen plotting position with
  mid-ranks for ties (Hyndman & Fan 1996 type 5, doi:10.1080/00031305.1996.10473566)
  as ``p = 100 * (r - 0.5) / n``. This module emits the same statistic as a
  fraction ``quality_rank = (r - 0.5) / n`` in (0, 1); the recipe's percentile
  is exactly ``100 * quality_rank`` (a fixed unit scaling — ordering and the
  §2.2 qualification thresholds are unaffected).
- **FERPA suppression** — ``*`` and blank cells parse to NaN (recipe §4.1);
  a school-vintage with a missing composite is emitted with
  ``metric_value = NaN`` and ``quality_rank = NaN`` and is EXCLUDED from the
  rank pool ``n`` for that vintage x level.
- **COVID rc2020** — no assessment was published (federal waiver, recipe
  §2.3); the loader detects the absent proficiency columns and emits
  carry-forward rows from rc2019 with ``stale=True`` covering the rc2020
  in-force window. Two consecutive assessment-less vintages would exceed the
  recipe's ``staleness_cap = 1`` and raise.
- **Pool caveat** — the recipe §2.1 rank pool is ultimately defined by the
  CPS attendance-boundary GeoJSON join (downstream of this module). Here the
  pool is the parse-stage approximation: CPS school-level rows classified to
  a level by the ISBE school-type field, with the recipe's belt-and-suspenders
  charter drop applied where a charter type exists (zip-era code ``C``;
  2018-2023 ``CHARTER SCH``). rc2024-rc2025 publish NO charter type (recipe
  §6 taxonomy-drift check), so charters remain in those two pools until the
  boundary join excludes them structurally — the recipe names the boundary
  join, not the type string, as the load-bearing exclusion. Rows with a
  missing school-type value (observed: 4 CPS rows in rc2019) cannot join a
  level pool and are dropped.
"""

from __future__ import annotations

import io
import re
import zipfile
from collections.abc import Iterable
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from lp_reit_lab.ingest.config import CPS_RCDTS_DISTRICT_PREFIX

# --- Era boundaries (recipe §1.2/§1.3 regime table; rc vintage years) -------
# ISAT/PSAE composite era ends rc2014; PARCC begins rc2015 (recipe §1.3).
_ISAT_PSAE_LAST_RC_YEAR = 2014
# HS sat PARCC in rc2015-rc2016 only; rc2017 HS switched to SAT (recipe §1.3,
# verified: rc2017 PARCC fields empty for Lincoln Park HS).
_PARCC_BOTH_LEVELS_LAST_RC_YEAR = 2016
# Regime labels break after rc2014 (ISAT/PSAE), rc2018 (PARCC incl. the
# PARCC/SAT split years), rc2024 (IAR+SAT); rc2025+ is IAR+ACT (recipe §1.2).
_PARCC_REGIME_LAST_RC_YEAR = 2018
_IAR_SAT_REGIME_LAST_RC_YEAR = 2024

# recipe §2.3: staleness cap = 1 missed vintage — equals the maximum observed
# publication gap in the 20-vintage series (the 2020 waiver is the only gap).
_STALENESS_CAP = 1

# recipe §3.2: RC vintage N is in force [Nov 1 of N, Oct 31 of N+1] —
# statutory Oct 31 preparation deadline (105 ILCS 5/10-17a) + verified
# late-October releases make Nov 1 the conservative knowability floor.
_IN_FORCE_START_MONTH, _IN_FORCE_START_DAY = 11, 1
_IN_FORCE_END_MONTH, _IN_FORCE_END_DAY = 10, 31

# --- Layout description strings (recipe §1.3 table + §6 verification log) ---
# Exact normalized descriptions of the zip-era metadata fields (verified
# identical rc2006-rc2017 in the official layout workbooks).
_DESC_RCDTS = "SCHOOL ID (R-C-D-T-S)"
_DESC_TYPE_CODE = "SCHOOL TYPE CODE (0,1,2,C)"
_DESC_SCHOOL_NAME = "SCHOOL NAME"
# rc2006-rc2014 composite: description prefix + group cell "ALL TESTS"
# disambiguates from the per-assessment (ISAT/PSAE/IAA/IMAGE) repeats and,
# via the embedded year, from the prior-year restatement fields the same file
# carries (recipe §1.3: never take restated history from a later vintage).
_ZIP_COMPOSITE_GROUP = "ALL TESTS"
_ZIP_COMPOSITE_DESC_FMT = "{year} SCHOOL COMPOSITE PERCENT FOR MEETS & EXCEEDS"
# rc2015-rc2017 PARCC levels 4+5 (recipe §1.3; exact match keeps the
# "PARTICIALLY/PARTIALLY MET EXPECTATIONS" level-3 fields out).
_PARCC_GROUP = "PARCC"
_PARCC_MET_DESC_FMT = "{year} ELA/MATH SCHOOL - STUDENTS MET EXPECTATIONS"
_PARCC_EXCEEDED_DESC_FMT = "{year} ELA/MATH SCHOOL - STUDENTS EXCEEDED EXPECTATIONS"
# rc2017 high-school SAT fields (recipe §1.3, fields #481/#482 verified).
_SAT_GROUP = "SAT"
_SAT_MEETS_DESC_FMT = "{year} ELA/MATH SCHOOL MEETS STANDARDS"
_SAT_EXCEEDS_DESC_FMT = "{year} ELA/MATH SCHOOL EXCEEDS STANDARDS"

# --- Workbook-era (rc2018+) sheet/column names (recipe §1.3 table) ----------
# Sheet located by name pattern: space-stripped uppercase equality, so
# "ELA and Math" (rc2018), "ELA Math Science" (rc2019-rc2023) and
# "ELAMathScience" (rc2024-rc2025) all resolve while "ELAMathScience (2)"
# (rc2025 subgroup continuation sheet) does not collide.
_WORKBOOK_SHEET_KEYS = frozenset({"ELAANDMATH", "ELAMATHSCIENCE"})
# Column-name pairs for the ELA/math school percentages; rc2018 uses the
# "Total %" naming, rc2019+ the "% ... Proficiency" naming (recipe §1.3,
# verified per year). A vintage whose sheet carries NEITHER pair is an
# assessment-less (COVID-waiver) vintage -> carry-forward (recipe §2.3).
_WORKBOOK_METRIC_PAIRS = (
    ("% ELA Proficiency", "% Math Proficiency"),
    ("ELA Proficiency Total %", "Math Proficiency Total %"),
)
_COL_RCDTS = "RCDTS"
# Row-granularity column: "Type" rc2018-rc2024, "Level" rc2025 (verified);
# school rows carry the value "School" in both.
_ROW_TYPE_COLS = ("Type", "Level")
_ROW_TYPE_SCHOOL = "School"
_COL_SCHOOL_NAME = "School Name"
_COL_SCHOOL_TYPE = "School Type"

# --- Level classification (recipe §2.1, §4.4; §6 school-type codes) ---------
# Zip-era SCHOOL TYPE CODE: 0=HIGH SCHOOL, 1=MIDDLE SCHL, 2=ELEMENTARY,
# C=CHARTER SCH (recipe §6, verified on the rc14 full-state file).
_ZIP_TYPE_CODE_TO_LEVEL = {"0": "high_school", "2": "elementary"}
# Codes structurally outside the two-level pool: middle schools (recipe §4.4)
# and the charter belt-and-suspenders drop (recipe §2.1).
_ZIP_TYPE_CODES_DROPPED = frozenset({"1", "C"})
# Workbook-era School Type strings (recipe §6 taxonomy-drift check:
# ELEMENTARY/HIGH SCHOOL/CHARTER SCH/MIDDLE SCHL/PreK through rc2023;
# Elementary School/High School/Middle/Junior High School/Special School
# from rc2024). Keys are whitespace-normalized uppercase.
_WORKBOOK_TYPE_TO_LEVEL = {
    "ELEMENTARY": "elementary",
    "ELEMENTARY SCHOOL": "elementary",
    "HIGH SCHOOL": "high_school",
}
_WORKBOOK_TYPES_DROPPED = frozenset(
    {
        "CHARTER SCH",  # recipe §2.1 belt-and-suspenders charter drop
        "MIDDLE SCHL",  # recipe §4.4 middle schools drop structurally
        "MIDDLE/JUNIOR HIGH SCHOOL",  # recipe §4.4 (rc2024+ naming)
        "PREK",  # recipe §2.1: no tested grades
        "SPECIAL SCHOOL",  # recipe §2.1: no attendance boundary (rc2024+)
    }
)

# RCDTS: 15-character region-county-district-type-school key (recipe §2.1);
# rc2025 publishes it dash-separated (15-016-2990-25-XXXX), earlier vintages
# compact. School suffix "0000" is a district/central-office aggregate.
_RCDTS_SCHOOL_SUFFIX_LEN = 4
_RCDTS_DISTRICT_SUFFIX = "0000"

# Zip-era rc{yy}.txt: semicolon-delimited, fields space-padded to layout
# widths (recipe §6). latin-1 decodes every byte 1:1, so school-name bytes
# outside ASCII never abort the stream; names are descriptive-only here.
_ZIP_TXT_ENCODING = "latin-1"
_ZIP_TXT_DELIMITER = ";"

# FERPA-suppressed cells publish as "*" (2018+ workbooks) or blank/space
# fields (zip era); both parse to missing (recipe §4.1).
_SUPPRESSED_SENTINEL = "*"

_RC_FILE_RE = re.compile(r"^rc(\d{4})[_-]", re.IGNORECASE)
_ZIP_LAYOUT_SHEET_RE = re.compile(r"RC\d{2}", re.IGNORECASE)
_ASSESSMENT_SHEET_KEY = "ASSESSMENT"

_OUTPUT_COLUMNS = [
    "rc_year",
    "rcdts",
    "school_name",
    "level",
    "metric_value",
    "quality_rank",
    "in_force_start",
    "in_force_end",
    "stale",
    "metric_regime",
]


def _norm(text: object) -> str:
    """Whitespace-normalized uppercase string for description/name matching."""
    return " ".join(str(text).split()).upper()


def _metric_regime(rc_year: int) -> str:
    """Assessment-regime label for an rc vintage (recipe §1.2)."""
    if rc_year <= _ISAT_PSAE_LAST_RC_YEAR:
        return "ISAT/PSAE"
    if rc_year <= _PARCC_REGIME_LAST_RC_YEAR:
        return "PARCC"
    if rc_year <= _IAR_SAT_REGIME_LAST_RC_YEAR:
        return "IAR+SAT"
    return "IAR+ACT"


def _in_force_window(rc_year: int) -> tuple[pd.Timestamp, pd.Timestamp]:
    """[Nov 1 of rc_year, Oct 31 of rc_year+1] per recipe §3.2."""
    start = pd.Timestamp(year=rc_year, month=_IN_FORCE_START_MONTH, day=_IN_FORCE_START_DAY)
    end = pd.Timestamp(year=rc_year + 1, month=_IN_FORCE_END_MONTH, day=_IN_FORCE_END_DAY)
    return start, end


def _normalize_rcdts(raw: object) -> str:
    """Canonical compact RCDTS: strip separators, uppercase (rc2025 is dashed)."""
    return re.sub(r"[^0-9A-Z]", "", str(raw).upper())


def _parse_pct(raw: object) -> float:
    """Parse a proficiency cell; '*' / blank -> NaN (recipe §4.1), else float."""
    if raw is None:
        return np.nan
    if isinstance(raw, int | float):
        return float(raw)
    text = str(raw).strip()
    if text in ("", _SUPPRESSED_SENTINEL):
        return np.nan
    return float(text)


def hazen_rank(values: pd.Series) -> pd.Series:
    """Hazen plotting-position rank fraction ``(r - 0.5) / n`` in (0, 1).

    ``r`` is the ascending mid-rank (ties averaged) and ``n`` the count of
    non-missing values; NaN inputs get NaN ranks and do not count toward
    ``n``. Recipe §1.2 (Hyndman & Fan 1996 type 5); the recipe's percentile
    is ``100 *`` this fraction.
    """
    ranks = values.rank(method="average")  # ascending mid-ranks; NaN -> NaN
    n = int(values.notna().sum())
    if n == 0:
        return pd.Series(np.nan, index=values.index, dtype=float)
    return (ranks - 0.5) / n


# --------------------------------------------------------------------------
# Raw-directory inventory
# --------------------------------------------------------------------------


def _inventory(raw_dir: Path) -> dict[int, dict[str, Path | None]]:
    """Map rc year -> {'data': path, 'layout': path|None} from file names.

    Layout files carry 'layout' in the name; a .zip data file marks the
    zip era, a non-layout .xlsx the workbook era.
    """
    vintages: dict[int, dict[str, Path | None]] = {}
    for path in sorted(Path(raw_dir).iterdir()):
        match = _RC_FILE_RE.match(path.name)
        if match is None:
            continue
        year = int(match.group(1))
        slot = vintages.setdefault(year, {"data": None, "layout": None})
        if "layout" in path.name.lower():
            slot["layout"] = path
        elif path.suffix.lower() in (".zip", ".xlsx"):
            slot["data"] = path
    return {year: slot for year, slot in vintages.items() if slot["data"] is not None}


# --------------------------------------------------------------------------
# Zip era (rc2006-rc2017): layout-resolved semicolon-delimited txt
# --------------------------------------------------------------------------


def _layout_rows(layout_path: Path, sheet: str) -> list[tuple[int, list[str]]]:
    """(field_number, normalized strings in the row) for each field row.

    A field row leads with the field number; section-header rows lead with
    prose and are skipped. The field number is the first numeric cell in row
    order (verified structure of the official layouts, recipe §1.3).
    """
    frame = pd.ExcelFile(layout_path).parse(sheet, header=None)
    rows: list[tuple[int, list[str]]] = []
    for _, row in frame.iterrows():
        number: int | None = None
        for cell in row.tolist():
            if isinstance(cell, bool):
                continue
            if isinstance(cell, int | float) and not pd.isna(cell):
                number = int(cell)
                break
            if isinstance(cell, str):
                stripped = cell.strip()
                if stripped.isdigit():
                    number = int(stripped)
                    break
                if stripped:  # prose before any number -> section header row
                    break
        if number is None:
            continue
        strings = [_norm(cell) for cell in row.tolist() if isinstance(cell, str) and cell.strip()]
        rows.append((number, strings))
    return rows


def _resolve_field(
    rows: list[tuple[int, list[str]]],
    description: str,
    group: str | None = None,
    prefix: bool = False,
) -> int:
    """0-based delimited-file index of the uniquely matching layout field.

    Match = normalized description equality (or prefix), optionally requiring
    ``group`` among the row's other cells (e.g. 'ALL TESTS' vs the ISAT/PSAE
    per-assessment repeats). Index = field number - 1 (recipe §1.3/§6).
    """
    target = _norm(description)
    hits = []
    for number, strings in rows:
        matched = any(s.startswith(target) if prefix else s == target for s in strings)
        if matched and (group is None or _norm(group) in strings):
            hits.append(number)
    if len(hits) != 1:
        raise ValueError(
            f"layout resolution for {description!r} (group={group!r}) "
            f"matched fields {hits}; expected exactly one"
        )
    return hits[0] - 1


def _zip_layout_sheet(layout_path: Path, data_member: str) -> str:
    """Layout sheet describing the zip's data member.

    rc2015-rc2017 ship '*assessment*' members described by the 'Assessment'
    sheet; rc2006-rc2014 members are described by the 'RC{yy}' sheet.
    """
    sheet_names = pd.ExcelFile(layout_path).sheet_names
    if "assessment" in data_member.lower():
        wanted = [s for s in sheet_names if _norm(s) == _ASSESSMENT_SHEET_KEY]
    else:
        wanted = [s for s in sheet_names if _ZIP_LAYOUT_SHEET_RE.fullmatch(s.strip())]
    if len(wanted) != 1:
        raise ValueError(
            f"{layout_path.name}: cannot pick layout sheet for member {data_member!r} "
            f"from {sheet_names}"
        )
    return wanted[0]


def _zip_metric_indices(
    rows: list[tuple[int, list[str]]], rc_year: int
) -> dict[str, tuple[int, ...]]:
    """Per-level tuples of 0-based indices whose values SUM to the metric.

    rc2006-rc2014: one ALL TESTS composite field, both levels (recipe §1.3).
    rc2015-rc2016: PARCC met + exceeded (levels 4+5), both levels.
    rc2017: PARCC met+exceeded for elementary, SAT meets+exceeds for HS.
    """
    if rc_year <= _ISAT_PSAE_LAST_RC_YEAR:
        composite = _resolve_field(
            rows,
            _ZIP_COMPOSITE_DESC_FMT.format(year=rc_year),
            group=_ZIP_COMPOSITE_GROUP,
            prefix=True,  # description tail names the subject mix, which drifts (recipe §1.3)
        )
        return {"elementary": (composite,), "high_school": (composite,)}
    parcc = (
        _resolve_field(rows, _PARCC_MET_DESC_FMT.format(year=rc_year), group=_PARCC_GROUP),
        _resolve_field(rows, _PARCC_EXCEEDED_DESC_FMT.format(year=rc_year), group=_PARCC_GROUP),
    )
    if rc_year <= _PARCC_BOTH_LEVELS_LAST_RC_YEAR:
        return {"elementary": parcc, "high_school": parcc}
    sat = (
        _resolve_field(rows, _SAT_MEETS_DESC_FMT.format(year=rc_year), group=_SAT_GROUP),
        _resolve_field(rows, _SAT_EXCEEDS_DESC_FMT.format(year=rc_year), group=_SAT_GROUP),
    )
    return {"elementary": parcc, "high_school": sat}


def _sum_or_nan(parts: list[float]) -> float:
    """Sum of the metric components; NaN if ANY component is suppressed.

    A partially suppressed composite is not the published signal, so it is
    treated as missing (recipe §4.1) rather than a silently smaller sum.
    """
    return float(np.sum(parts)) if not any(np.isnan(p) for p in parts) else np.nan


def _parse_zip_vintage(rc_year: int, data_path: Path, layout_path: Path | None) -> pd.DataFrame:
    """CPS school rows (rcdts, school_name, level, metric_value) for a zip vintage."""
    if layout_path is None:
        raise ValueError(f"rc{rc_year}: zip-era vintage requires a layout workbook")
    with zipfile.ZipFile(data_path) as archive:
        members = [m for m in archive.namelist() if m.lower().endswith(".txt")]
        if len(members) != 1:
            raise ValueError(f"{data_path.name}: expected exactly one .txt member, got {members}")
        member = members[0]
        sheet = _zip_layout_sheet(layout_path, member)
        rows = _layout_rows(layout_path, sheet)
        idx_rcdts = _resolve_field(rows, _DESC_RCDTS)
        idx_type = _resolve_field(rows, _DESC_TYPE_CODE)
        idx_name = _resolve_field(rows, _DESC_SCHOOL_NAME)
        metric_indices = _zip_metric_indices(rows, rc_year)

        records: list[dict[str, object]] = []
        with archive.open(member) as handle:  # in-memory stream; no disk extraction
            for line in io.TextIOWrapper(handle, encoding=_ZIP_TXT_ENCODING):
                # RCDTS is field 1 -> the line starts with it (verified).
                if not line.startswith(CPS_RCDTS_DISTRICT_PREFIX):
                    continue
                fields = line.rstrip("\r\n").split(_ZIP_TXT_DELIMITER)
                rcdts = _normalize_rcdts(fields[idx_rcdts])
                if rcdts[-_RCDTS_SCHOOL_SUFFIX_LEN:] == _RCDTS_DISTRICT_SUFFIX:
                    continue  # district/central-office aggregate, not a school
                type_code = fields[idx_type].strip()
                if type_code in _ZIP_TYPE_CODES_DROPPED:
                    continue  # charter (§2.1) / middle school (§4.4)
                level = _ZIP_TYPE_CODE_TO_LEVEL.get(type_code)
                if level is None:
                    raise ValueError(
                        f"rc{rc_year}: unknown school type code {type_code!r} for "
                        f"RCDTS {rcdts} — re-verify against the recipe §6 code table"
                    )
                metric = _sum_or_nan([_parse_pct(fields[i]) for i in metric_indices[level]])
                records.append(
                    {
                        "rcdts": rcdts,
                        "school_name": fields[idx_name].strip(),
                        "level": level,
                        "metric_value": metric,
                    }
                )
    return pd.DataFrame.from_records(records)


# --------------------------------------------------------------------------
# Workbook era (rc2018+): multi-sheet xlsx public data sets
# --------------------------------------------------------------------------


def _workbook_sheet_name(sheet_names: list[str]) -> str:
    """The unique ELA/math proficiency sheet (name pattern varies by year)."""
    wanted = [s for s in sheet_names if _norm(s).replace(" ", "") in _WORKBOOK_SHEET_KEYS]
    if len(wanted) != 1:
        raise ValueError(f"cannot locate the ELA/math sheet among {sheet_names}; got {wanted}")
    return wanted[0]


def _header_index(header: tuple[object, ...]) -> tuple[dict[str, int], set[str]]:
    """(name -> first position, set of duplicated names).

    Real workbooks carry duplicated names among UNUSED columns (verified:
    rc2018 repeats 'Math Participation Total IEP Count'), so duplication only
    fails loud when a column this parser actually consumes is ambiguous.
    """
    index: dict[str, int] = {}
    duplicated: set[str] = set()
    for position, cell in enumerate(header):
        if cell is None:
            continue
        name = str(cell).strip()
        if not name:
            continue
        if name in index:
            duplicated.add(name)
        else:
            index[name] = position
    return index, duplicated


def _workbook_school_record(
    rc_year: int,
    row: tuple[object, ...],
    header: dict[str, int],
    metric_cols: tuple[str, str],
) -> dict[str, object] | None:
    """One tidy record for a CPS school row; None for rows outside the pool."""
    rcdts = _normalize_rcdts(row[header[_COL_RCDTS]])
    if not rcdts.startswith(CPS_RCDTS_DISTRICT_PREFIX):
        return None
    school_type = row[header[_COL_SCHOOL_TYPE]]
    type_key = _norm(school_type) if school_type is not None else ""
    if not type_key or type_key in _WORKBOOK_TYPES_DROPPED:
        # No school-type value -> cannot join a level pool; dropped types per
        # recipe §2.1 (charter belt-and-suspenders) and §4.4 (middle/PreK).
        return None
    level = _WORKBOOK_TYPE_TO_LEVEL.get(type_key)
    if level is None:
        raise ValueError(
            f"rc{rc_year}: unknown School Type {school_type!r} — taxonomy "
            "drift; re-verify against the recipe §6 log before mapping"
        )
    subject_values = [_parse_pct(row[header[c]]) for c in metric_cols]
    metric = (
        float(np.mean(subject_values))
        if not any(np.isnan(v) for v in subject_values)
        else np.nan
    )
    return {
        "rcdts": rcdts,
        "school_name": str(row[header[_COL_SCHOOL_NAME]] or "").strip(),
        "level": level,
        "metric_value": metric,
    }


def _parse_workbook_vintage(rc_year: int, data_path: Path) -> pd.DataFrame | None:
    """CPS school rows for a workbook vintage; None if no assessment published.

    Metric = unweighted mean of the ELA and math school percentages (recipe
    §1.3: mirrors the pooled composite ISBE itself published 2006-2017);
    missing if EITHER subject is suppressed (recipe §4.1).
    """
    workbook = load_workbook(data_path, read_only=True)
    try:
        sheet = _workbook_sheet_name(workbook.sheetnames)
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        header: dict[str, int] | None = None
        duplicated: set[str] = set()
        for row in rows:  # header = first row carrying the RCDTS key column
            if any(isinstance(c, str) and c.strip() == _COL_RCDTS for c in row):
                header, duplicated = _header_index(row)
                break
        if header is None:
            raise ValueError(f"sheet {sheet!r}: no header row containing {_COL_RCDTS!r}")

        metric_cols = next(
            (pair for pair in _WORKBOOK_METRIC_PAIRS if all(c in header for c in pair)),
            None,
        )
        if metric_cols is None:
            return None  # assessment-less vintage (COVID waiver, recipe §2.3)

        row_type_col = next((c for c in _ROW_TYPE_COLS if c in header), None)
        if row_type_col is None:
            raise ValueError(f"sheet {sheet!r}: no {_ROW_TYPE_COLS} row-granularity column")
        for required in (_COL_SCHOOL_NAME, _COL_SCHOOL_TYPE):
            if required not in header:
                raise ValueError(f"sheet {sheet!r}: required column {required!r} missing")
        used = {*metric_cols, row_type_col, _COL_RCDTS, _COL_SCHOOL_NAME, _COL_SCHOOL_TYPE}
        ambiguous = used & duplicated
        if ambiguous:
            raise ValueError(f"sheet {sheet!r}: duplicated header(s) {sorted(ambiguous)}")

        records = [
            record
            for row in rows
            if row[header[row_type_col]] == _ROW_TYPE_SCHOOL  # drop aggregates
            if (record := _workbook_school_record(rc_year, row, header, metric_cols))
            is not None
        ]
    finally:
        workbook.close()
    return pd.DataFrame.from_records(records)


# --------------------------------------------------------------------------
# Assembly
# --------------------------------------------------------------------------


def _parse_vintage(
    rc_year: int, slot: dict[str, Path | None]
) -> pd.DataFrame | None:
    data_path = slot["data"]
    assert data_path is not None  # _inventory() only keeps vintages with data
    if data_path.suffix.lower() == ".zip":
        frame = _parse_zip_vintage(rc_year, data_path, slot["layout"])
    else:
        frame = _parse_workbook_vintage(rc_year, data_path)
    if frame is None:
        return None
    if frame.empty:
        raise ValueError(f"rc{rc_year}: no CPS school rows parsed — wrong or corrupt input file")
    if frame["rcdts"].duplicated().any():
        duplicated = frame.loc[frame["rcdts"].duplicated(), "rcdts"].tolist()
        raise ValueError(f"rc{rc_year}: duplicate RCDTS keys {duplicated} would corrupt ranks")
    return frame


def _ranked(frame: pd.DataFrame) -> pd.DataFrame:
    """Attach quality_rank within each level pool (recipe §1.2/§2.1)."""
    out = frame.copy()
    out["quality_rank"] = out.groupby("level")["metric_value"].transform(hazen_rank)
    return out


def _finalize(frame: pd.DataFrame, rc_year: int, stale: bool, regime: str) -> pd.DataFrame:
    out = frame.copy()
    start, end = _in_force_window(rc_year)
    out["rc_year"] = rc_year
    out["in_force_start"] = start
    out["in_force_end"] = end
    out["stale"] = stale
    out["metric_regime"] = regime
    return out[_OUTPUT_COLUMNS]


def load_school_quality(
    raw_dir: Path, years: Iterable[int] | None = None
) -> pd.DataFrame:
    """Load the tidy CPS school-quality frame from an ISBE raw snapshot dir.

    Parameters
    ----------
    raw_dir:
        Snapshot directory holding ``rc{year}_...`` data files (zips with
        their layout workbooks for rc2006-rc2017, public-data-set xlsx for
        rc2018+), e.g. ``data/raw/isbe_report_card/snapshot=2026-07-06``.
    years:
        Optional subset of rc vintage years to emit; default all discovered.
        An assessment-less vintage (rc2020) pulls its carry-forward source
        (the prior vintage) automatically even when not requested.

    Returns
    -------
    One row per CPS school x vintage with columns
    ``rc_year, rcdts, school_name, level, metric_value, quality_rank,
    in_force_start, in_force_end, stale, metric_regime`` (see module
    docstring for semantics). Sorted by (rc_year, level, rcdts).
    """
    inventory = _inventory(Path(raw_dir))
    if not inventory:
        raise ValueError(f"{Path(raw_dir).as_posix()}: no rc{{year}}_* data files found")
    if years is None:
        requested = sorted(inventory)
    else:
        requested = sorted(set(years))
        missing = [y for y in requested if y not in inventory]
        if missing:
            raise ValueError(f"requested rc years {missing} not present in {raw_dir}")

    parsed: dict[int, pd.DataFrame | None] = {}

    def parse(year: int) -> pd.DataFrame | None:
        if year not in parsed:
            frame = _parse_vintage(year, inventory[year])
            parsed[year] = None if frame is None else _ranked(frame)
        return parsed[year]

    pieces: list[pd.DataFrame] = []
    for year in requested:
        ranked = parse(year)
        if ranked is not None:
            pieces.append(_finalize(ranked, year, stale=False, regime=_metric_regime(year)))
            continue
        # Assessment-less vintage: carry forward the last published ranks
        # with a staleness flag (recipe §2.3). staleness_cap = 1 means the
        # source must itself be a published (non-carried) vintage.
        source_year = year - _STALENESS_CAP
        if source_year not in inventory:
            raise ValueError(
                f"rc{year}: no assessment published and no rc{source_year} "
                "carry-forward source in the snapshot (recipe §2.3)"
            )
        source = parse(source_year)
        if source is None:
            raise ValueError(
                f"rc{year} and rc{source_year} both publish no assessment — "
                f"exceeds staleness_cap={_STALENESS_CAP} (recipe §2.3)"
            )
        pieces.append(
            _finalize(source, year, stale=True, regime=_metric_regime(source_year))
        )

    result = pd.concat(pieces, ignore_index=True)
    return result.sort_values(["rc_year", "level", "rcdts"], ignore_index=True)
